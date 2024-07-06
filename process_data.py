from openpyxl import load_workbook
import os

def xlsx_to_json(file: str) -> dict:
	""" Function to convert XLSX file to Python Dict """
	if os.path.isfile(file):
		workbook_object = load_workbook(file, data_only=True)
	else:
		raise FileNotFoundError(f"File {file} is not found!")
	
	workbook_object = load_workbook(file, data_only=True)
	sheets = [workbook_object[sheetname] for sheetname in workbook_object.sheetnames]
	# We assume that different sheets may have different datasets of the same type
	data = {}

	for worksheet in sheets:
		data[worksheet.title] = {}
		for row in range(3, worksheet.max_row + 1):
			key = int(worksheet.cell(row, 1).value.replace(" ", ""))
			# Columns 2 and 3 (name and personal information) are anonymized
			values = {
				"Institute":                    worksheet.cell(row, 4).value,
				"Group":                        worksheet.cell(row, 5).value,
				"Jurisprudence":                worksheet.cell(row, 6).value,
				"Physical Training":            worksheet.cell(row, 7).value,
				"Engineering Graphics Light":   worksheet.cell(row, 8).value,
				"History Light":                worksheet.cell(row, 9).value,
				"Life Safety":                  worksheet.cell(row, 10).value,
				"Engineering Graphics Middle":  worksheet.cell(row, 11).value,
				"History Middle":               worksheet.cell(row, 12).value,
				"Chemistry":                    worksheet.cell(row, 13).value,
				"English Middle":               worksheet.cell(row, 14).value,
				"Geodesy":                      worksheet.cell(row, 15).value,
				"Engineering Graphics Hard":    worksheet.cell(row, 16).value,
				"Computer science":             worksheet.cell(row, 17).value,
				"Programming":                  worksheet.cell(row, 18).value,
				"Practical Phonetics":          worksheet.cell(row, 19).value,
				"Mathematics":                  worksheet.cell(row, 20).value,
				"English Hard":                 worksheet.cell(row, 21).value,
				"Academic Failures":            int(worksheet.cell(row, 22).value),
				"Status":                       worksheet.cell(row, 23).value,
			}
			data[worksheet.title][key] = values

	return data
